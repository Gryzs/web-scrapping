import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

headers = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
}

url = 'http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid=36482&module=M'
site = requests.get(url, headers=headers)
status = site.status_code

if status == 200:
    soup = BeautifulSoup(site.content, 'html.parser')
    # Encontrar a tabela pelo ID 'grd_DXMainTable'
    table = soup.find('table', id='grd_DXMainTable')

    if table:
        # Inicializar uma lista para armazenar os dados
        data = []

        # Encontrar todas as linhas da tabela (tr)
        rows = table.find_all('tr')

        # Iterar sobre as linhas (começando da primeira linha de dados)
        for row in rows[1:]:  # Começa do índice 1 para pular a linha de cabeçalho
            # Encontrar todas as células da linha (td)
            cells = row.find_all('td')

            # Extrair o texto de cada célula e adicionar à lista de dados
            row_data = [cell.get_text(strip=True) for cell in cells]
            data.append(row_data)

        # Criar um novo arquivo Excel e adicionar os dados
        wb = Workbook()
        ws = wb.active
        ws.title = 'Dados da Tabela'

        # Colocar o cabeçalho na planilha (supondo que o cabeçalho está na primeira linha da lista 'data')
        header = data[0]
        for col_index, header_value in enumerate(header):
            ws.cell(row=1, column=col_index+1, value=header_value)

        # Escrever os dados na planilha a partir da segunda linha
        for row_index, row_data in enumerate(data[1:], start=2):  # Começa do índice 2 para pular o cabeçalho
            for col_index, cell_value in enumerate(row_data):
                ws.cell(row=row_index, column=col_index+1, value=cell_value)

        # Salvar o arquivo Excel
        excel_file = 'dados_tabela.xlsx'
        wb.save(excel_file)
        print(f"Dados salvos com sucesso em '{excel_file}'.")

    else:
        print("Tabela não encontrada com o ID especificado.")
else:
    print(f"Erro ao acessar o site. Status code: {status}")
